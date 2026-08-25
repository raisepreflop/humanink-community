#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys, os

CARPETA = sys.argv[1] if len(sys.argv)>1 else '.'
SEMANAS = int(sys.argv[2]) if len(sys.argv)>2 else 4

OUTPUT = Path(CARPETA) / 'social'
OUTPUT.mkdir(exist_ok=True)

wb = openpyxl.Workbook()

# ── Colors ──────────────────────────────────────────────────────────────
C = {
    'ig_reel':     'E1306C', 'ig_carrusel': 'C13584', 'ig_story': 'FCAF45',
    'fb_post':     '4267B2', 'fb_grupo':    '1877F2', 'fb_video': '3b5998',
    'yt_video':    'FF0000', 'yt_short':    'FF4444',
    'tofu':        'E8F5E9', 'mofu':        'FFF3E0', 'bofu':     'FCE4EC',
    'header':      '1a1a2e', 'subheader':   'e94560',
    'white':       'FFFFFF', 'light':       'F5F5F5', 'border':   'DDDDDD',
}

def fill(hex_color): return PatternFill(fill_type='solid', fgColor=hex_color)
def font(bold=False, color='111111', size=11):
    return Font(bold=bold, color=color, size=size)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)
def border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

# ═══════════════════════════════════════════════════════════════
# SHEET 1 — WEEKLY VIEW
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Weekly Calendar'

DIAS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
CANALES = [
    ('Instagram', 'Reel', 'ig_reel'),
    ('Instagram', 'Carousel/Feed', 'ig_carrusel'),
    ('Instagram', 'Story', 'ig_story'),
    ('Facebook', 'Page post', 'fb_post'),
    ('Facebook', 'Group', 'fb_grupo'),
    ('YouTube', 'Video/Short', 'yt_video'),
]

# Main header
ws1.merge_cells('A1:I1')
ws1['A1'] = '📅 CONTENT CALENDAR — SOCIAL NETWORKS'
ws1['A1'].fill = fill(C['header'])
ws1['A1'].font = font(bold=True, color='FFFFFF', size=14)
ws1['A1'].alignment = center()
ws1.row_dimensions[1].height = 30

# Temperature colors for the calendar
TEMP_COLORS = {
    'TOFU (Cold)':   'E8F5E9',
    'MOFU (Warm)':  'FFF3E0',
    'BOFU (Hot)': 'FCE4EC',
}

for sem in range(1, SEMANAS + 1):
    row_start = 2 + (sem - 1) * (len(CANALES) + 3)

    # Week header
    ws1.merge_cells(f'A{row_start}:I{row_start}')
    ws1[f'A{row_start}'] = f'WEEK {sem}'
    ws1[f'A{row_start}'].fill = fill(C['subheader'])
    ws1[f'A{row_start}'].font = font(bold=True, color='FFFFFF', size=12)
    ws1[f'A{row_start}'].alignment = center()
    ws1.row_dimensions[row_start].height = 22

    # Days header
    ws1[f'A{row_start+1}'] = 'Channel'
    ws1[f'B{row_start+1}'] = 'Format'
    ws1[f'A{row_start+1}'].fill = fill('333333')
    ws1[f'B{row_start+1}'].fill = fill('333333')
    ws1[f'A{row_start+1}'].font = font(bold=True, color='FFFFFF', size=9)
    ws1[f'B{row_start+1}'].font = font(bold=True, color='FFFFFF', size=9)
    ws1[f'A{row_start+1}'].alignment = center()
    ws1[f'B{row_start+1}'].alignment = center()

    for d, dia in enumerate(DIAS):
        col = get_column_letter(d + 3)
        cell = ws1[f'{col}{row_start+1}']
        cell.value = dia
        cell.fill = fill('333333')
        cell.font = font(bold=True, color='FFFFFF', size=9)
        cell.alignment = center()

    # Channel rows
    for c_i, (canal, formato, color_key) in enumerate(CANALES):
        row = row_start + 2 + c_i
        ws1[f'A{row}'] = canal
        ws1[f'B{row}'] = formato
        ws1[f'A{row}'].fill = fill(C[color_key] + '44' if len(C[color_key])==6 else C[color_key])
        ws1[f'B{row}'].fill = fill(C[color_key] + '22' if len(C[color_key])==6 else C[color_key])
        ws1[f'A{row}'].font = font(bold=True, size=9)
        ws1[f'B{row}'].font = font(size=9)
        ws1[f'A{row}'].alignment = center()
        ws1[f'B{row}'].alignment = center()
        ws1.row_dimensions[row].height = 50

        for d in range(7):
            col = get_column_letter(d + 3)
            cell = ws1[f'{col}{row}']
            cell.fill = fill(C['light'])
            cell.font = font(size=8)
            cell.alignment = left()
            cell.border = border()

    ws1.row_dimensions[row_start+1].height = 18

# Widths
ws1.column_dimensions['A'].width = 13
ws1.column_dimensions['B'].width = 16
for d in range(7):
    ws1.column_dimensions[get_column_letter(d+3)].width = 22

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — DETAILED LIST OF PIECES
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Content List')

headers2 = [
    'Week', 'Day', 'Network', 'Format', 'Size (px)',
    'Pillar', 'Temperature', 'Hook (1st line)',
    'Message / Core idea', 'CTA', 'Destination',
    'Hashtags', 'Publish time', 'Status', 'Notes'
]
ws2.merge_cells('A1:O1')
ws2['A1'] = '📋 DETAILED CONTENT LIST'
ws2['A1'].fill = fill(C['header'])
ws2['A1'].font = font(bold=True, color='FFFFFF', size=13)
ws2['A1'].alignment = center()
ws2.row_dimensions[1].height = 28

for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.fill = fill(C['subheader'])
    cell.font = font(bold=True, color='FFFFFF', size=9)
    cell.alignment = center()
    cell.border = border()
ws2.row_dimensions[2].height = 18

# Example content for the first 2 weeks
EJEMPLOS = [
    [1,'Mon','Instagram','Reel 9:16','1080×1920','Craft','TOFU','Why did your favorite book hook you in the first line?','The literary hook technique: the 3 openings that always work','Save it if you want to use it 👇','Link in bio','#writing #writer #reading','10:00','Scheduled','Repurpose → YouTube Short'],
    [1,'Tue','Instagram','Carousel 1:1','1080×1080','Book','TOFU','This character was born from a real nightmare','5 facts about [Character] you didn\'t know','Swipe to the end →','Preorder in bio','#novel #characters #thriller','12:00','Draft','Include photo of author\'s notes'],
    [1,'Wed','Facebook','Page post','1200×630','Author','MOFU','It took me [N] years to publish my first book. Here\'s what I learned.','Personal story + lesson applicable for writers','How long have you been working on your book? Tell me 👇','Group in comments','','14:00','To create','Link in first comment'],
    [1,'Thu','YouTube','Short 9:16','1080×1920','Craft','TOFU','The 3-word trick to write dialogue that sounds real','Quick dialogue technique: verb + emotion + subtext','Subscribe for more techniques','YouTube channel','#writing #dialogue','15:00','To create','Repurpose from IG Reel'],
    [1,'Fri','Instagram','Reel 9:16','1080×1920','Community','MOFU','How many pages do you read a day?','Mini poll + personal result + author\'s reading habit','Reply with your number in comments →','WhatsApp community in bio','#reading #habits #readers','11:00','To create','Boost for the weekend'],
    [1,'Sat','Instagram','Story','1080×1920','Promotion','BOFU','This weekend: chapter 1 free 👀','First chapter teaser · Direct link','Link in the story →','Tally / Substack','','11:30','To create','Link sticker'],
    [2,'Mon','Instagram','Carousel 1:1','1080×1080','Craft','TOFU','7 books every [genre] writer should read','Curated list with 1 explanatory line per book','Which one have you read? Comment the number 👇','Link in bio','#books #recommendations','10:00','To create','Authority pillar'],
    [2,'Tue','Facebook','Group','1200×630','Community','MOFU','Debate: Can good writing be taught or is it just talent?','Genuine debate post · Share your personal opinion','What do you think? 👇','Group','','12:00','To create','Group only, not page'],
]

for r, row_data in enumerate(EJEMPLOS, 3):
    temp = row_data[6]
    row_color = C['tofu'] if temp=='TOFU' else C['mofu'] if temp=='MOFU' else C['bofu']
    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=col, value=val)
        cell.fill = fill(row_color)
        cell.font = font(size=8)
        cell.alignment = left()
        cell.border = border()

# Empty rows to fill in
for r in range(len(EJEMPLOS)+3, len(EJEMPLOS)+3 + (SEMANAS-2)*7 + 20):
    for col in range(1, len(headers2)+1):
        cell = ws2.cell(row=r, column=col)
        cell.fill = fill(C['white'] if r%2==0 else C['light'])
        cell.border = border()
        if col == 13: cell.value = '—'  # Time

# Column widths sheet 2
anchos2 = [7,7,12,16,12,12,14,35,50,25,18,25,12,12,20]
for i, w2 in enumerate(anchos2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w2
ws2.freeze_panes = 'A3'

# ═══════════════════════════════════════════════════════════════
# SHEET 3 — WEEKLY KPI TRACKER
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('KPI Tracker')

ws3.merge_cells('A1:L1')
ws3['A1'] = '📊 WEEKLY KPI TRACKER'
ws3['A1'].fill = fill(C['header'])
ws3['A1'].font = font(bold=True, color='FFFFFF', size=13)
ws3['A1'].alignment = center()
ws3.row_dimensions[1].height = 28

kpi_headers = [
    'Week', 'Network', 'Reach', 'Engagement Rate %',
    'Saves / Shares', 'Link clicks', 'New followers',
    'Conv. Substack', 'Conv. WhatsApp', 'Conv. Tally',
    'Attributable sales', 'Notes / Key insight'
]
for col, h in enumerate(kpi_headers, 1):
    cell = ws3.cell(row=2, column=col, value=h)
    cell.fill = fill(C['subheader'])
    cell.font = font(bold=True, color='FFFFFF', size=9)
    cell.alignment = center()
    cell.border = border()

for sem in range(1, SEMANAS*2+1):
    for red in ['Instagram', 'Facebook', 'YouTube']:
        row = 2 + (sem-1)*3 + ['Instagram','Facebook','YouTube'].index(red) + 1
        cell = ws3.cell(row=row, column=1, value=f'Week {sem}')
        cell.font = font(bold=True, size=9)
        cell.fill = fill(C['light'])
        cell.border = border()
        cell2 = ws3.cell(row=row, column=2, value=red)
        cell2.font = font(size=9)
        cell2.fill = fill(C['light'])
        cell2.border = border()
        for col in range(3, len(kpi_headers)+1):
            c = ws3.cell(row=row, column=col)
            c.border = border()
            c.fill = fill(C['white'])

for i, w3 in enumerate([8,14,12,16,20,14,16,15,15,14,16,40], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w3
ws3.freeze_panes = 'A3'

# ═══════════════════════════════════════════════════════════════
# SHEET 4 — HOOK BANK
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Hook Bank')

ws4.merge_cells('A1:E1')
ws4['A1'] = '🪝 HOOK BANK — Hook-Message-CTA Structure'
ws4['A1'].fill = fill(C['header'])
ws4['A1'].font = font(bold=True, color='FFFFFF', size=13)
ws4['A1'].alignment = center()
ws4.row_dimensions[1].height = 28

hook_headers = ['Hook type', 'Hook (first 2 lines)', 'Emotion it triggers', 'Best format', 'Temperature']
for col, h in enumerate(hook_headers, 1):
    cell = ws4.cell(row=2, column=col, value=h)
    cell.fill = fill(C['subheader'])
    cell.font = font(bold=True, color='FFFFFF', size=9)
    cell.alignment = center()
    cell.border = border()

HOOKS_BANCO = [
    ['Uncomfortable question', 'Why do most writers never finish their novel?\n(And how you can be the exception)', 'Curiosity + identity', 'Reel / Carousel cover', 'TOFU'],
    ['Controversial statement', '80% of writing advice is garbage.\nHere\'s what actually works.', 'Outrage + curiosity', 'Reel', 'TOFU'],
    ['Personal story', '12 publishers rejected me.\nToday I have [N] published books.', 'Empathy + hope', 'Reel / FB Post', 'TOFU-MOFU'],
    ['Surprising fact', 'The average reader decides in 7 seconds whether to buy a book.\nThis influences that decision.', 'Surprise + urgency', 'Carousel', 'TOFU'],
    ['Counterintuitive', 'Writing faster makes your novels better.\nNot the other way around.', 'Curiosity + bewilderment', 'Reel', 'TOFU'],
    ['Mistake revelation', 'I made the mistake every new writer makes.\nAnd I caught it on page 300.', 'Empathy + learning', 'FB Post / Reel', 'MOFU'],
    ['Numbered list', '5 books that changed how I write.\n(No one mentions the 3rd)', 'FOMO + curiosity', 'Carousel', 'TOFU'],
    ['Testimonial', '"I read your [N] books in one week."\nSomeone told me this this morning.', 'Social proof + emotion', 'Story / Post', 'MOFU-BOFU'],
    ['Behind the scenes', 'This is my desk at 6am.\nAnd what I have on screen right now.', 'Intimacy + connection', 'Story / Reel', 'MOFU'],
    ['Direct CTA', 'My new book is now available.\nYou can read the first chapter free here → 👇', 'Urgency + opportunity', 'Story / Post', 'BOFU'],
]

for r, row_data in enumerate(HOOKS_BANCO, 3):
    temp = row_data[4]
    row_color = C['tofu'] if 'TOFU' in temp else C['mofu'] if 'MOFU' in temp else C['bofu']
    for col, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=col, value=val)
        cell.fill = fill(row_color)
        cell.font = font(size=8)
        cell.alignment = left()
        cell.border = border()
    ws4.row_dimensions[r].height = 40

for i, w4 in enumerate([18, 55, 25, 20, 14], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w4

# ── Save ──────────────────────────────────────────────────────────────────
out = OUTPUT / 'calendario-contenidos.xlsx'
wb.save(str(out))
print(f"✓ Calendar: {out}")
print(f"  Sheets: Weekly Calendar · Content List · KPI Tracker · Hook Bank")
