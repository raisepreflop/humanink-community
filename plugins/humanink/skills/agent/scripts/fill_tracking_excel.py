import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

EXCEL_OUT = sys.argv[1]
TITULO    = sys.argv[2] if len(sys.argv) > 2 else ""

wb = openpyxl.load_workbook(EXCEL_OUT)
ws = wb["Tracking"]

DATA_FONT = Font(size=10, name="Calibri")
ALT_FILL  = PatternFill("solid", fgColor="EBF0FA")
thin = Side(style='thin', color="B8CCE4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# The Literary Agent replaces this list with the publishers for the project's genre,
# ordered by priority (High = best catalog match + accepts new authors)
# Columns: (Publisher, Group, Genres, Contact type, Email/URL, Priority)
# HumanInk COMPLETES THIS LIST based on biblia.md and the market analysis
ROWS = [
    # PLACEHOLDER — replace with the publishers suitable for the genre
]

# HumanInk: before writing the rows, read the project's genre
# and select the 10-15 most suitable publishers from the internal database
# of the Market Analyst (humanink:analyst). Order by priority.

for i, row in enumerate(ROWS):
    r = i + 4  # starts at row 4
    fill = ALT_FILL if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    valores = list(row) + ["—", "", "", "—", "—", "—", "—", ""]
    for c, val in enumerate(valores, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = DATA_FONT
        cell.border = border
        cell.fill = fill
        cell.alignment = LEFT if c in (1,2,3,5,13) else CENTER

# Update title with the project's real data
ws["A1"] = f"PUBLISHER TRACKING — {TITULO}"

wb.save(EXCEL_OUT)
print(f"✓ Excel filled with {len(ROWS)} publishers: {EXCEL_OUT}")
